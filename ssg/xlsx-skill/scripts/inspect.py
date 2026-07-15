"""Explore an .xlsx workbook: list sheets, headers, and dimensions.

Usage: python scripts/inspect.py <INPUT_PATH>
"""
import sys
import openpyxl


def inspect(path: str) -> None:
    wb = openpyxl.load_workbook(path, data_only=False)
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"[sheet] {name}  dims={ws.dimensions}  rows={ws.max_row} cols={ws.max_column}")
        header = [c.value for c in next(ws.iter_rows(max_row=1))]
        print(f"  header row 1: {header}")


if __name__ == "__main__":
    inspect(sys.argv[1] if len(sys.argv) > 1 else "input.xlsx")
